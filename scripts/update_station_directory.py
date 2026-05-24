#!/usr/bin/env python3
"""Download and parse the ÖBB station directory Excel file.

The script exports a simplified JSON mapping (``bst_id``, ``bst_code``, ``name``,
``in_vienna`` and ``pendler``) that is used throughout the project. Station names
are harmonized with the previous export, geodata is used to flag Vienna
locations and commuter belt entries, and stations outside this area are omitted.
The data is obtained from the official ÖBB Open-Data portal.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import logging
import os
import re

# Bandit B404: subprocess is required to invoke internal cache-refresh
# scripts. Inputs are static lists, never user-supplied.
import subprocess  # nosec B404
import sys
import unicodedata
from copy import deepcopy
from dataclasses import dataclass, field
import zipfile
from io import BytesIO
from pathlib import Path
from typing import cast
from collections.abc import Callable, Iterable, Mapping, MutableMapping, Sequence

import openpyxl

__all__ = ["subprocess"]

DEFAULT_SOURCE_URL = "https://data.oebb.at/dam/jcr:fce22daf-0dd8-4a15-80b4-dbca6e80ce38/" "Verzeichnis%20der%20Verkehrsstationen.xlsx"
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

try:  # pragma: no cover - convenience for module execution
    from src.utils.files import (
        atomic_write,
        loads_finite,
        read_capped_bytes,
        read_capped_json,
        read_capped_text,
        validate_zip_archive_safe,
    )
    from src.utils.geo import (
        apply_coordinate_inertia,
        use_cached_polygon_result,
    )
    from src.utils.http import fetch_content_safe, session_with_retries
    from src.utils.serialize import scrub_trojan_source_primitives
    from src.utils.stations import is_in_vienna as _is_point_in_vienna
except ModuleNotFoundError:  # pragma: no cover - fallback when installed as package
    from utils.files import (  # type: ignore[no-redef]
        atomic_write,
        loads_finite,
        read_capped_bytes,
        read_capped_json,
        read_capped_text,
        validate_zip_archive_safe,
    )
    from utils.geo import (  # type: ignore[no-redef]
        apply_coordinate_inertia,
        use_cached_polygon_result,
    )
    from utils.http import fetch_content_safe, session_with_retries  # type: ignore[no-redef]
    from utils.serialize import scrub_trojan_source_primitives  # type: ignore[no-redef]
    from utils.stations import is_in_vienna as _is_point_in_vienna  # type: ignore[no-redef]

# Security cap against wide-but-flat JSON size-bomb attacks. Mirrors the
# canonical ``MAX_*_FILE_BYTES`` contract from ``src/utils/cache.py`` /
# ``src/utils/stations.py``: depth-bomb catch alone misses ``MemoryError``
# (a ``BaseException`` subclass) so a planted-huge file (~1 GiB of
# ``[0,0,…]``) buffered via ``path.read_text()`` propagates past the
# loader and crashes the cron pipeline (the orchestrator runs every
# update script via ``subprocess.run(check=True)``). 50 MiB is ~285x
# the production stations.json so legitimate state is never rejected.
MAX_JSON_FILE_BYTES = 50 * 1024 * 1024

# Security cap against wide-but-flat CSV size-bomb attacks. Routes every
# operator-controlled CSV file (GTFS stops, WL haltepunkte, VOR
# haltestellen) through ``read_capped_text`` -> ``io.StringIO`` ->
# ``csv.DictReader`` so a planted unbounded CSV (single huge line, no
# newlines) cannot buffer GiB of payload via ``handle.readline()`` and
# propagate ``MemoryError`` (``BaseException`` subclass NOT caught by
# ``except (OSError, csv.Error)``) past the cron orchestrator.
# Same 50 MiB ceiling as ``MAX_JSON_FILE_BYTES`` — comfortably above any
# legitimate transit-network CSV (Austria-wide GTFS dumps stay well
# under 50 MiB) and well below the runner's 1 GiB cgroup limit.
MAX_CSV_LOCATIONS_BYTES = 50 * 1024 * 1024

# Security cap against planted-huge binary cache payloads at the ÖBB
# workbook fallback path (:data:`DEFAULT_CACHED_WORKBOOK_PATH`). The
# fallback branch in :func:`download_workbook` is reached when the live
# ``data.oebb.at`` fetch fails; ``read_capped_bytes`` is the canonical
# defence shape (mirrors :data:`MAX_JSON_FILE_BYTES` /
# :data:`MAX_CSV_LOCATIONS_BYTES` in this module and
# :data:`MAX_LOG_PRUNE_FILE_BYTES` in :mod:`src.feed.logging`). Pinned
# at 10 MiB — identical to :data:`src.utils.http.MAX_PAYLOAD_SIZE`, the
# upper bound the HTTP-fetch path could have legitimately produced — so
# a cache file larger than what HTTP could have stored is by definition
# tampered (compromised CI runner / hostile PR / manual operator dump /
# partial flush + power loss) and is rejected at the read boundary.
# Production xlsx (~62 KiB) is ~169x under the cap, so legitimate state
# is never rejected. Pre-fix: ``cache_path.read_bytes()`` allocates
# O(file_size) bytes before the surrounding ``except`` (none) runs, so
# a 10 GiB planted file at the cache path raises ``MemoryError`` past
# the surrounding cron orchestrator (``subprocess.run(check=True)`` in
# :mod:`scripts.update_all_stations`), aborting the WHOLE weekly cron
# tick. Post-fix the file is treated as missing and the original
# upstream ``Exception`` re-raised — mirrors the pre-fix shape on a
# missing cache file.
MAX_CACHED_WORKBOOK_BYTES = 10 * 1024 * 1024

try:  # pragma: no cover - convenience for module execution
    from src.places.client import (
        DEFAULT_INCLUDED_TYPES,
        GooglePlacesClient,
        GooglePlacesConfig,
        GooglePlacesError,
        GooglePlacesPermissionError,
        GooglePlacesTileError,
        Place,
        get_places_api_key,
    )
    from src.places.diagnostics import permission_hint
    from src.places.hafas_client import enrich_station_with_hafas
    from src.places.merge import BoundingBox, MergeConfig, merge_places, StationEntry
    from src.places.osm_client import (
        OSMOverpassError,
        VIENNA_BOUNDING_BOX,
        fetch_osm_places,
        filter_complete_places,
    )
    from src.places.tiling import Tile, iter_tiles, load_tiles_from_env, load_tiles_from_file
    from src.utils.env import get_bool_env, load_default_env_files
    from src.utils.logging import sanitize_log_arg
except ModuleNotFoundError:  # pragma: no cover - fallback when installed as package
    from places.client import (  # type: ignore[no-redef]
        DEFAULT_INCLUDED_TYPES,
        GooglePlacesClient,
        GooglePlacesConfig,
        GooglePlacesError,
        GooglePlacesPermissionError,
        GooglePlacesTileError,
        Place,
        get_places_api_key,
    )
    from places.diagnostics import permission_hint  # type: ignore[no-redef]
    from places.hafas_client import enrich_station_with_hafas  # type: ignore[no-redef]
    from places.merge import BoundingBox, MergeConfig, merge_places, StationEntry  # type: ignore[no-redef]
    from places.osm_client import (  # type: ignore[no-redef]
        OSMOverpassError,
        VIENNA_BOUNDING_BOX,
        fetch_osm_places,
        filter_complete_places,
    )
    from places.tiling import Tile, iter_tiles, load_tiles_from_env, load_tiles_from_file  # type: ignore[no-redef]
    from utils.env import get_bool_env, load_default_env_files  # type: ignore[no-redef]
    from utils.logging import sanitize_log_arg  # type: ignore[no-redef]

DEFAULT_OUTPUT_PATH = _ROOT / "data" / "stations.json"
DEFAULT_PENDLER_PATH = _ROOT / "data" / "pendler_bst_ids.json"
DEFAULT_PENDLER_CANDIDATES_PATH = _ROOT / "data" / "pendler_candidates.json"
DEFAULT_GTFS_STOPS_PATH = _ROOT / "data" / "gtfs" / "stops.txt"
DEFAULT_WL_HALTEPUNKTE_PATH = _ROOT / "data" / "wienerlinien-ogd-haltepunkte.csv"
DEFAULT_VOR_STOPS_PATH = _ROOT / "data" / "vor-haltestellen.csv"
# Compact projection of the ÖBB-Infrastruktur GeoNetz dataset (1 056 rail
# stations with their UIC EVA-Nr, IFOPT-ID and postal address). Generated
# by ``scripts/extract_oebb_geonetz_stops.py`` from the upstream 23 MiB
# ``GeoNetz_12-2024.zip``. Read by ``_enrich_with_geonetz`` to attach
# the three identifier fields onto every ÖBB-Excel station whose
# ``bst_id`` matches a row's ``bsts_id``. Coordinates are left untouched
# by this enrichment — they are governed by OSM/HAFAS/Google tiers and
# the PR #1601 GeoNetz-reconciliation; this loader only adds metadata.
DEFAULT_GEONETZ_STOPS_PATH = _ROOT / "data" / "oebb_geonetz_stops.json"
# Soft-fail snapshot for the ÖBB workbook — mirrors the pinned-CSV
# fallback pattern used by WL OGD (PR #1441-#1442) and the VOR
# haltestellen snapshot, so a transient ``data.oebb.at`` outage no
# longer zeroes-out a whole weekly cron tick. On every successful
# download the workbook bytes are atomically written here; on the
# next download failure ``download_workbook`` falls back to this
# cached copy (with a warning) before re-raising.
DEFAULT_CACHED_WORKBOOK_PATH = _ROOT / "data" / "oebb-verkehrsstationen.xlsx"
REQUEST_TIMEOUT = 30  # seconds
USER_AGENT = "wien-oepnv station updater " "(https://github.com/Origamihase/wien-oepnv)"

HEADER_VARIANTS: dict[str, set[str]] = {
    "name": {"verkehrsstation"},
    "bst_code": {"bstcode"},
    "bst_id": {"bstid"},
}

logger = logging.getLogger("update_station_directory")


def _path_fingerprint(path: Path) -> str:
    """Return a one-way SHA-256 fingerprint of ``str(path)`` (12 hex chars).

    Security (Path-Log Sibling Drift Round 2, ``scripts/`` closure):
    mirrors :func:`src.utils.env._path_fingerprint`. The path arguments
    at every caller-side WARNING / INFO log line in this script come
    from operator-controlled CLI flags (``--output``, ``--pendler``,
    ``--pendler-candidates``, ``--vor-stops``, ``--gtfs-stops``,
    ``--wl-haltepunkte``). Interpolating the raw path bytes lets a
    hostile path carrying Trojan-Source primitives (BiDi RLO,
    zero-width, 8-bit C1 CSI/OSC, Tag block, Variation Selectors,
    newline log-forgery, ANSI ESC) flow verbatim into stderr /
    aggregated cron logs / SIEM splitters. The hex-only fingerprint
    is Trojan-Source-clean and a CodeQL-recognised barrier for the
    ``py/clear-text-logging-sensitive-data`` taint.
    """
    return hashlib.sha256(
        str(path).encode("utf-8", errors="replace")
    ).hexdigest()[:12]


def _empty_args() -> tuple[str, ...]:
    return ()


@dataclass(frozen=True)
class CacheRefreshTarget:
    label: str
    script_candidates: tuple[str, ...]
    optional: bool = False
    extra_args_factory: Callable[[], Sequence[str]] = _empty_args
    availability_check: Callable[[], bool] | None = None


def _has_google_places_credentials() -> bool:
    return bool(os.getenv("GOOGLE_ACCESS_ID") or os.getenv("GOOGLE_MAPS_API_KEY"))


def _google_cache_args() -> tuple[str, ...]:
    return ("--write",)


_CACHE_REFRESH_TARGETS: tuple[CacheRefreshTarget, ...] = (
    CacheRefreshTarget("ÖBB", ("update_oebb_cache.py",)),
    # VOR was removed from the station-directory refresh path on
    # 2026-05-11 per operator policy: VOR is now used only for the
    # Stammstrecke delay monitor (every 30 min, via update-cycle.yml).
    # The station directory uses the pinned ``data/vor-haltestellen.csv``
    # snapshot for VOR stop IDs.
    CacheRefreshTarget("Wiener Linien", ("update_wl_cache.py",)),
    CacheRefreshTarget(
        "Google Places",
        ("update_google_cache.py", "fetch_google_places_stations.py"),
        optional=True,
        extra_args_factory=_google_cache_args,
        availability_check=_has_google_places_credentials,
    ),
)


def _refresh_provider_caches(*, script_dir: Path | None = None) -> None:
    base_dir = script_dir or Path(__file__).resolve().parent
    for target in _CACHE_REFRESH_TARGETS:
        if target.availability_check and not target.availability_check():
            logger.info("Skipping %s cache refresh (credentials not available)", target.label)
            continue

        command: list[str] | None = None
        script_path: Path | None = None
        for candidate in target.script_candidates:
            candidate_path = base_dir / candidate
            if candidate_path.exists():
                extra_args = list(target.extra_args_factory())
                command = [sys.executable, str(candidate_path), *extra_args]
                script_path = candidate_path
                break

        if command is None or script_path is None:
            message = "No cache refresh script found for %s; skipping" if target.optional else "Cache refresh script missing for %s"
            log_method = logger.debug if target.optional else logger.warning
            log_method(message, target.label)
            continue

        logger.info("Refreshing %s cache via %s", target.label, script_path.name)
        try:
            # Enforce a 5-minute timeout to prevent indefinite hangs (DoS protection)
            result = subprocess.run(command, check=False, shell=False, timeout=300)  # nosec B603
        except subprocess.TimeoutExpired:
            logger.warning("%s cache refresh timed out after 300s; continuing", target.label)
            continue
        except OSError as exc:  # pragma: no cover - execution environment issues
            logger.warning("Failed to execute %s cache refresh (%s); continuing", target.label, exc)
            continue

        if result.returncode != 0:
            logger.warning(
                "%s cache refresh exited with code %s; continuing",
                target.label,
                result.returncode,
            )
        else:
            logger.info("%s cache refresh completed", target.label)


@dataclass
class Station:
    """Representation of a single station entry."""

    bst_id: str
    bst_code: str
    name: str
    in_vienna: bool = False
    pendler: bool = False
    vor_id: str | None = None
    extras: dict[str, object] = field(default_factory=dict)

    def as_dict(self) -> dict[str, object]:
        payload: dict[str, object] = {
            "bst_id": str(self.bst_id),
            "bst_code": self.bst_code,
            "name": self.name,
            "in_vienna": self.in_vienna,
            "pendler": self.pendler,
        }
        if self.vor_id:
            payload["vor_id"] = self.vor_id

        for key, value in self.extras.items():
            if key in {"bst_id", "bst_code", "name", "in_vienna", "pendler", "vor_id", "_lat", "_lng", "_types", "_google_place_id"}:
                continue
            payload[key] = value
        # Tag every Excel-imported entry with source="oebb" if no other
        # source has been merged in. Without this, the next run treats the
        # entry as `manual` (see _load_existing_station_entries) and appends
        # it on top of a fresh Excel import — producing the duplicate-name
        # NamingIssues seen in the 2026-05 cron after PR #1201.
        if "source" not in payload:
            payload["source"] = "oebb"
        return payload

    def update_from_entry(self, entry: Mapping[str, object]) -> None:
        base_keys = {"bst_id", "bst_code", "name", "in_vienna", "pendler", "vor_id"}

        # Coordinate Inertia: capture the existing coords BEFORE the
        # generic extras-loop (below) overwrites them. We use these
        # later to decide whether to absorb upstream drift via
        # ``apply_coordinate_inertia``.
        existing_lat_raw = self.extras.get("latitude")
        existing_lon_raw = self.extras.get("longitude")
        existing_lat = (
            float(existing_lat_raw)
            if isinstance(existing_lat_raw, int | float)
            else None
        )
        existing_lon = (
            float(existing_lon_raw)
            if isinstance(existing_lon_raw, int | float)
            else None
        )

        for key, value in entry.items():
            if key == "vor_id":
                if self.vor_id is None and isinstance(value, str) and value.strip():
                    self.vor_id = value.strip()
                continue
            if key in base_keys:
                continue
            self.extras[key] = deepcopy(value)

        lat = entry.get("latitude") or entry.get("_lat")
        lng = entry.get("longitude") or entry.get("_lng")
        new_lat = float(lat) if isinstance(lat, int | float) else None
        new_lon = float(lng) if isinstance(lng, int | float) else None

        # Resolve the final coordinate via the inertia helper. Drifts
        # below ``STATION_DRIFT_TOLERANCE_METERS`` (default 150 m) are
        # absorbed — the existing coords win, eliminating churn-only
        # diffs in ``data/stations.json``. Drifts at or above the
        # threshold are accepted as legitimate relocations. First-time
        # coords (no existing) and missing-from-upstream cases are
        # handled by the helper's rules 1+2.
        merged_lat, merged_lon = apply_coordinate_inertia(
            existing_lat, existing_lon, new_lat, new_lon
        )
        if merged_lat is not None:
            self.extras["latitude"] = merged_lat
        if merged_lon is not None:
            self.extras["longitude"] = merged_lon


@dataclass
class VORStop:
    """Minimal representation of a VOR stop for ID matching."""

    vor_id: str
    name: str
    municipality: str | None = None
    short_name: str | None = None


@dataclass
class LocationInfo:
    """Coordinates collected from auxiliary data sources."""

    latitude: float
    longitude: float
    sources: set[str]

    def add_source(self, source: str) -> None:
        self.sources.add(source)


def _strip_accents(value: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch))


def _normalize_location_keys(name: str | None) -> list[str]:
    if not name:
        return []

    candidates = [name]
    without_parens = re.sub(r"\s*\([^)]*\)\s*", " ", name)
    if without_parens not in candidates:
        candidates.append(without_parens)
    stripped_suffix = re.sub(
        r"\b(?:Bahnsteig|Bahnsteige|Gleis)\b[^,;/]*",
        " ",
        without_parens,
        flags=re.IGNORECASE,
    )
    if stripped_suffix not in candidates:
        candidates.append(stripped_suffix)

    tokens: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        token = _strip_accents(candidate)
        token = token.replace("ß", "ss")
        token = token.casefold()
        token = re.sub(r"\b(?:bahnhof|bahnhst|bhf|hbf|bf)\b", "", token)
        token = re.sub(r"\b(?:bahnsteig|bahnsteige|gleis)\b", "", token)
        token = token.replace("-", " ").replace("/", " ")
        token = re.sub(r"[^a-z0-9]+", " ", token)
        token = re.sub(r"\s{2,}", " ", token).strip()
        if token and token not in seen:
            seen.add(token)
            tokens.append(token)
        numeric_stripped = re.sub(r"\d+", "", token).strip()
        if numeric_stripped and numeric_stripped not in seen:
            seen.add(numeric_stripped)
            tokens.append(numeric_stripped)
    return tokens


def _harmonize_station_name(name: str) -> str:
    """Return *name* with unified whitespace for consistent lookups."""

    text = "".join(" " if ch in {"\u00a0", "\u2007", "\u202f"} else ch for ch in str(name))
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip()


def _coerce_float_value(value: object | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _store_location(
    locations: dict[str, LocationInfo],
    key: str,
    lat: float,
    lon: float,
    source: str,
) -> None:
    info = locations.get(key)
    if info is None:
        locations[key] = LocationInfo(latitude=lat, longitude=lon, sources={source})
        return
    info.add_source(source)
    # Prefer coordinates from station level data when replacing placeholder zeros
    if (info.latitude, info.longitude) == (0.0, 0.0):
        info.latitude = lat
        info.longitude = lon


def _load_gtfs_locations(path: Path) -> dict[str, LocationInfo]:
    locations: dict[str, LocationInfo] = {}
    if not path.exists():
        logger.warning(
            "GTFS stops file not found: [path-sha256=%s]",
            _path_fingerprint(path),
        )
        return locations
    # Security: route through ``read_capped_text`` to bound the
    # ``csv.DictReader`` -> ``readline()`` allocation against planted
    # unbounded CSVs. See ``MAX_CSV_LOCATIONS_BYTES`` for the threat
    # model. ``read_capped_text`` returns ``None`` on missing /
    # oversized / decode-error so subsequent code receives an empty
    # mapping rather than crashing the cron pipeline.
    content = read_capped_text(
        path,
        MAX_CSV_LOCATIONS_BYTES,
        encoding="utf-8",
        label="GTFS stops",
        logger=logger,
    )
    if content is None:
        return locations
    try:
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            stop_name = row.get("stop_name")
            if not stop_name:
                continue
            stop_name = _harmonize_station_name(stop_name)
            lat = _coerce_float_value(row.get("stop_lat"))
            lon = _coerce_float_value(row.get("stop_lon"))
            if lat is None or lon is None:
                continue
            location_type = (row.get("location_type") or "").strip()
            is_station = location_type == "1"
            for key in _normalize_location_keys(stop_name):
                if not key:
                    continue
                if is_station or key not in locations:
                    _store_location(locations, key, lat, lon, source="gtfs")
    except csv.Error as exc:
        logger.warning(
            "Could not parse GTFS stops file [path-sha256=%s]: %s",
            _path_fingerprint(path),
            exc,
        )
    else:
        if locations:
            logger.info("Loaded %d GTFS stop coordinates", len(locations))
    return locations


_WL_DIVA_KEY_PREFIX = "diva:"


def _wl_diva_key(diva: str) -> str:
    """Return the namespaced location-index key for a WL DIVA identifier.

    ``_normalize_location_keys`` strips ``:`` (and casefolds), so a DIVA
    key can never collide with a name-derived key — letting callers that
    hold a station's ``wl_diva`` resolve coordinates by the authoritative
    identifier instead of the lossy ``Wien <name> (WL)`` <-> ``<name>``
    name match.
    """

    return f"{_WL_DIVA_KEY_PREFIX}{diva}"


def _load_wienerlinien_locations(path: Path) -> dict[str, LocationInfo]:
    locations: dict[str, LocationInfo] = {}
    if not path.exists():
        logger.warning(
            "Wiener Linien haltepunkte file not found: [path-sha256=%s]",
            _path_fingerprint(path),
        )
        return locations
    # Security: see _load_gtfs_locations for the canonical CSV
    # size-bomb defence shape (read_capped_text + io.StringIO).
    content = read_capped_text(
        path,
        MAX_CSV_LOCATIONS_BYTES,
        encoding="utf-8",
        label="WL haltepunkte",
        logger=logger,
    )
    if content is None:
        return locations
    try:
        reader = csv.DictReader(io.StringIO(content), delimiter=";")
        for row in reader:
            # The Wiener Linien OGD CSV schema migrated (PR #1442): the
            # legacy data.wien.gv.at proxy export keyed on NAME /
            # WGS84_LAT / WGS84_LON; the canonical wienerlinien.at
            # OGD-Echtzeit export that replaced it renames those to
            # StopText / Latitude / Longitude and exposes the DIVA. Read
            # the new names first and fall back to the legacy ones so the
            # loader survives either upstream shape (mirrors the fuzzy-key
            # resilience in scripts/update_wl_stations.py).
            name = row.get("StopText") or row.get("NAME")
            if name:
                name = _harmonize_station_name(name)
            lat = _coerce_float_value(row.get("Latitude") or row.get("WGS84_LAT"))
            lon = _coerce_float_value(row.get("Longitude") or row.get("WGS84_LON"))
            if lat is None or lon is None:
                continue
            diva = (row.get("DIVA") or "").strip()
            if diva:
                _store_location(locations, _wl_diva_key(diva), lat, lon, source="wl")
            if not name:
                continue
            for key in _normalize_location_keys(name):
                if not key:
                    continue
                _store_location(locations, key, lat, lon, source="wl")
    except csv.Error as exc:
        logger.warning(
            "Could not parse Wiener Linien haltepunkte file [path-sha256=%s]: %s",
            _path_fingerprint(path),
            exc,
        )
    else:
        if locations:
            logger.info("Loaded %d Wiener Linien coordinates", len(locations))
    return locations


def _load_vor_locations(path: Path) -> dict[str, LocationInfo]:
    """Load name → (lat, lon) from the VOR haltestellen CSV.

    Provides a third coordinate source after GTFS and WL haltepunkte.
    The VOR CSV uses ``StopPointName;Latitude;Longitude`` columns.
    """
    locations: dict[str, LocationInfo] = {}
    if not path.exists():
        logger.warning(
            "VOR stops file not found: [path-sha256=%s]",
            _path_fingerprint(path),
        )
        return locations
    # Security: see _load_gtfs_locations for the canonical CSV
    # size-bomb defence shape (read_capped_text + io.StringIO).
    content = read_capped_text(
        path,
        MAX_CSV_LOCATIONS_BYTES,
        encoding="utf-8-sig",
        label="VOR stops",
        logger=logger,
    )
    if content is None:
        return locations
    try:
        sample = content[:4096]
        delimiter = _detect_csv_delimiter(sample)
        reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
        for row in reader:
            name = row.get("StopPointName") or row.get("Name") or row.get("StopName")
            if name:
                name = _harmonize_station_name(name)
            lat = _coerce_float_value(row.get("Latitude") or row.get("WGS84_LAT"))
            lon = _coerce_float_value(row.get("Longitude") or row.get("WGS84_LON"))
            if not name or lat is None or lon is None:
                continue
            for key in _normalize_location_keys(name):
                if not key:
                    continue
                _store_location(locations, key, lat, lon, source="vor")
    except csv.Error as exc:
        logger.warning(
            "Could not parse VOR stops file [path-sha256=%s]: %s",
            _path_fingerprint(path),
            exc,
        )
    else:
        if locations:
            logger.info("Loaded %d VOR coordinates", len(locations))
    return locations


def _build_location_index(
    gtfs_path: Path | None,
    wl_path: Path | None,
    vor_path: Path | None = None,
) -> dict[str, LocationInfo]:
    locations: dict[str, LocationInfo] = {}
    if gtfs_path:
        locations.update(_load_gtfs_locations(gtfs_path))
    if wl_path:
        wl_locations = _load_wienerlinien_locations(wl_path)
        for key, value in wl_locations.items():
            _store_location(locations, key, value.latitude, value.longitude, source="wl")
    if vor_path:
        vor_locations = _load_vor_locations(vor_path)
        for key, value in vor_locations.items():
            # GTFS/WL already populated entries take precedence (more authoritative
            # for transit-platform coords); VOR fills gaps for stations that are
            # neither in the local GTFS snapshot nor in the WL haltepunkte CSV.
            if key not in locations:
                _store_location(locations, key, value.latitude, value.longitude, source="vor")
    return locations


def _load_existing_station_entries(
    path: Path,
) -> tuple[dict[str, dict[str, object]], list[dict[str, object]]]:
    if not path.exists():
        return {}, []
    # Security: ``read_capped_json`` enforces both the depth-bomb catch
    # tuple and the byte-size cap (see MAX_JSON_FILE_BYTES). Without
    # the cap a wide-but-flat planted file would propagate
    # ``MemoryError`` past the loader and crash the cron pipeline.
    payload = read_capped_json(
        path,
        MAX_JSON_FILE_BYTES,
        label="Existing station directory",
    )
    if payload is None:
        logger.warning(
            "Could not parse existing station directory "
            "[path-sha256=%s] (missing/invalid/oversized)",
            _path_fingerprint(path),
        )
        return {}, []

    if isinstance(payload, dict):
        payload = payload.get("stations", [])

    mapping: dict[str, dict[str, object]] = {}
    manual_stations: list[dict[str, object]] = []

    if isinstance(payload, list):
        for entry in payload:
            if not isinstance(entry, dict):
                continue

            bst_id = entry.get("bst_id")
            source = entry.get("source", "")

            is_manual = False
            if bst_id is None:
                is_manual = True
            elif source != "oebb":
                if isinstance(source, str):
                    stripped = source.strip()
                    if stripped:
                        # Token-based classification: a substring match for
                        # ``"oebb"`` inside a comma-separated source string
                        # would mis-classify entries whose ``source`` carries
                        # only a metadata provider whose name *contains* the
                        # canonical ``oebb`` token (e.g. ``oebb_geonetz`` —
                        # the GeoNetz EVA/IFOPT enrichment, not the ÖBB
                        # Excel ``Verzeichnis der Verkehrsstationen``).
                        # Pre-fix the substring check at this site bucketed
                        # the synthetic ``Wien Hauptbahnhof`` (``bst_id=
                        # 900100``, ``source=google_places,oebb_geonetz,
                        # vor,wl``) and ``Wien Kaiserebersdorf`` (``bst_id=
                        # 900105``, ``source=oebb_geonetz,vor``) into the
                        # ÖBB-Excel ``mapping`` even though the underlying
                        # ``bst_id`` is never present in the live ÖBB
                        # workbook. The entries therefore failed to round-
                        # trip through the rebuild and dropped out of
                        # ``data/stations.json`` on every weekly cron tick —
                        # the regression visible in commit ``484c1f6``'s
                        # post-mortem (the file fix there did not address
                        # the underlying classifier bug).
                        tokens = {t.strip() for t in stripped.split(",") if t.strip()}
                        is_manual = "oebb" not in tokens
                    else:
                        # Backward-compat: entries written before the
                        # ``as_dict`` source-default fix lack a source
                        # field entirely (empty string). If they carry the
                        # typical ÖBB Excel fields (``bst_id`` +
                        # ``bst_code``), treat them as ÖBB — otherwise the
                        # next Excel pull would create a duplicate and
                        # trip the canonical-name uniqueness gate (see
                        # PR #1203 cron failure post-mortem).
                        bst_code = entry.get("bst_code")
                        is_manual = not (isinstance(bst_code, str) and bst_code.strip())
                elif isinstance(source, list) and "oebb" in source:
                    is_manual = False
                elif not source:
                    # Backward-compat (non-string variant): ``None`` / unset
                    # ``source`` falls through to the same ``bst_code`` guard.
                    bst_code = entry.get("bst_code")
                    is_manual = not (isinstance(bst_code, str) and bst_code.strip())
                else:
                    is_manual = True

            if is_manual:
                manual_stations.append(entry)
                continue

            if bst_id is not None:
                try:
                    lookup_id = str(int(float(bst_id)))
                except (ValueError, TypeError):
                    lookup_id = str(bst_id)
                if lookup_id:
                    mapping[lookup_id] = entry

    return mapping, manual_stations


def _restore_existing_metadata(stations: Iterable[Station], existing_entries: dict[str, dict[str, object]]) -> None:
    for station in stations:
        try:
            lookup_id = str(int(float(station.bst_id)))
        except (ValueError, TypeError):
            lookup_id = str(station.bst_id)
        existing = existing_entries.get(lookup_id)
        if not existing:
            continue
        vor_id_raw = existing.get("vor_id")
        if isinstance(vor_id_raw, str):
            vor_id = vor_id_raw.strip()
            if vor_id:
                station.vor_id = vor_id
        for key, value in existing.items():
            if key in {"bst_id", "bst_code", "name", "in_vienna", "pendler", "vor_id"}:
                continue
            station.extras[key] = deepcopy(value)


def _looks_like_vienna(text: str | None) -> bool:
    if not text:
        return False
    normalized = text.strip().casefold()
    if not normalized.startswith("wien"):
        return False
    if len(normalized) == 4:
        return True
    return not normalized[4].isalpha()


def _detect_csv_delimiter(sample: str) -> str:
    semicolons = sample.count(";")
    commas = sample.count(",")
    if semicolons >= commas and semicolons:
        return ";"
    if commas:
        return ","
    return ";"


def _normalize_csv_key(value: str | None) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def _parse_included_types(raw: str | None) -> list[str]:
    if raw is None:
        return list(DEFAULT_INCLUDED_TYPES)
    items = [part.strip() for part in raw.split(",") if part.strip()]
    return items or list(DEFAULT_INCLUDED_TYPES)


def _parse_radius(raw: str | None) -> int:
    if raw is None:
        return 2500
    try:
        radius = int(raw)
    except (TypeError, ValueError):
        # Security (Path-Log Sibling Drift Round 4, env-repr closure):
        # ``raw`` is the operator-controlled ``PLACES_RADIUS_M`` value.
        # Pre-fix the WARNING line interpolated it via ``%r`` — Python's
        # repr() escapes most attack bytes but lets all 256 Variation
        # Selectors (U+FE00-U+FE0F + U+E0100-U+E01EF) through verbatim
        # into ``record.args[0]`` and ``record.getMessage()``. Route
        # through ``sanitize_log_arg`` so the canonical
        # ``_INVISIBLE_DANGEROUS_RE`` strips them BEFORE the value
        # lands in caplog / non-SafeFormatter handlers.
        logger.warning(
            "Invalid PLACES_RADIUS_M=%s – using default 2500",
            sanitize_log_arg(raw),
        )
        return 2500
    return max(1, min(50000, radius))


def _parse_max_results(raw: str | None) -> int:
    if raw is None:
        return 20
    try:
        value = int(raw)
    except (TypeError, ValueError):
        # Security: see ``_parse_radius`` — same env-repr drift closure.
        logger.warning(
            "Invalid PLACES_MAX_RESULTS=%s – using default 20",
            sanitize_log_arg(raw),
        )
        return 20
    if value <= 0:
        return 0
    return max(1, min(20, value))


def _parse_float(raw: str | None, *, key: str, default: float) -> float:
    if raw is None:
        return default
    try:
        return float(raw)
    except (TypeError, ValueError):
        # Security: see ``_parse_radius`` — same env-repr drift closure.
        # ``key`` is a hardcoded module-internal constant
        # (``REQUEST_TIMEOUT_S`` / ``MERGE_MAX_DIST_M``); only ``raw`` is
        # operator-controlled and therefore the only arg requiring the
        # canonical scrub.
        logger.warning(
            "Invalid %s=%s – using default %s",
            key,
            sanitize_log_arg(raw),
            default,
        )
        return default


def _parse_int(raw: str | None, *, key: str, default: int) -> int:
    if raw is None:
        return default
    try:
        return int(raw)
    except (TypeError, ValueError):
        # Security: see ``_parse_radius`` — same env-repr drift closure.
        logger.warning(
            "Invalid %s=%s – using default %s",
            key,
            sanitize_log_arg(raw),
            default,
        )
        return default


def _parse_bounding_box(raw: str | None) -> BoundingBox | None:
    if not raw:
        return None
    # Security: ``RecursionError`` covers JSON depth-bomb attacks in the
    # ``BOUNDINGBOX_VIENNA`` env override (intentional misconfig, leaked
    # CI env, compromised secret store). ``json.loads`` raises
    # ``RecursionError`` (NOT a subclass of ``json.JSONDecodeError`` and
    # NOT caught by the outer ``except ValueError`` in
    # ``_enrich_with_google_places``) on a deeply-nested payload. Without
    # this catch the unhandled error crashes the entire
    # ``update_station_directory.py`` cron pipeline.
    try:
        # Security: ``loads_finite`` pins parse_constant + parse_float
        # hooks (Round 1503 sibling) that reject NaN / Infinity / 1e1000
        # literals planted into the env-controlled ``BOUNDINGBOX_VIENNA``
        # override (leaked CI env / compromised secret store / hostile
        # operator). Mirrors the canonical defence pinned at
        # ``scripts/fetch_google_places_stations.py:_parse_bounding_box``.
        data = loads_finite(raw)
    except (json.JSONDecodeError, RecursionError) as exc:
        raise ValueError("BOUNDINGBOX_VIENNA must be valid JSON") from exc
    if not isinstance(data, dict):
        raise ValueError("BOUNDINGBOX_VIENNA must define min_lat/min_lng/max_lat/max_lng")
    try:
        return BoundingBox(
            min_lat=float(data["min_lat"]),
            min_lng=float(data["min_lng"]),
            max_lat=float(data["max_lat"]),
            max_lng=float(data["max_lng"]),
        )
    except (KeyError, TypeError, ValueError) as exc:
        raise ValueError("BOUNDINGBOX_VIENNA must define min_lat/min_lng/max_lat/max_lng") from exc


def _load_tiles_configuration(tiles_file: Path | None, env: MutableMapping[str, str]) -> Sequence[Tile]:
    if tiles_file:
        return load_tiles_from_file(tiles_file)
    return load_tiles_from_env(env.get("PLACES_TILES"))


def _fetch_google_places(client: GooglePlacesClient, tiles: Sequence[Tile]) -> list[Place]:
    places_by_id: dict[str, Place] = {}
    for tile in iter_tiles(tiles):
        logger.info("Fetching Google Places tile")
        try:
            for place in client.iter_nearby([tile]):
                places_by_id.setdefault(place.place_id, place)
        except GooglePlacesTileError as exc:
            logger.warning(
                "Skipping Google Places tile due to error: %s",
                exc,
            )
            continue
    return list(places_by_id.values())


def _merge_google_metadata(
    stations: list[Station],
    places: Sequence[Place],
    merge_config: MergeConfig,
) -> None:
    if not places:
        logger.info("Google Places enrichment returned no places")
        return

    existing_entries = cast(list[StationEntry], [station.as_dict() for station in stations])
    outcome = merge_places(existing_entries, places, merge_config)

    by_id: dict[str, Mapping[str, object]] = {}
    for entry in outcome.stations:
        bst_id = entry.get("bst_id")
        if isinstance(bst_id, str) and bst_id:
            by_id[str(bst_id)] = entry

    for station in stations:
        merged = by_id.get(station.bst_id)
        if merged:
            station.update_from_entry(merged)

    logger.info(
        "Google Places enrichment updated %d stations; %d suggestions already covered",
        len(outcome.updated_entries),
        len(outcome.skipped_places),
    )

    unmatched = [entry for entry in outcome.new_entries if "bst_id" not in entry]
    if unmatched:
        logger.info(
            "Google Places suggested %d additional stations without bst_id; ignoring",
            len(unmatched),
        )


def _stations_missing_coordinates(stations: list[Station]) -> list[Station]:
    """Return the subset of *stations* that still lack lat/lng metadata.

    The OSM-first / Google-fallback split keys on this set: OSM is the
    primary directory enrichment source; Google Places is only invoked
    if at least one station is still missing coordinates after the OSM
    merge has run. Without this gate every cron tick would burn the
    Google Places monthly free-tier quota even when OSM already covered
    the entire directory.
    """
    missing: list[Station] = []
    for station in stations:
        lat = station.extras.get("latitude")
        lng = station.extras.get("longitude")
        if not isinstance(lat, int | float) or not isinstance(lng, int | float):
            missing.append(station)
    return missing


def _enrich_with_osm(
    stations: list[Station],
    *,
    bounding_box: BoundingBox | None,
    merge_distance_m: float,
) -> bool:
    """Enrich *stations* via the OSM Overpass API.

    Returns ``True`` if the upstream call succeeded (regardless of how
    many stations actually got updated), ``False`` if OSM failed and
    the caller should fall through to the Google Places path. The
    bounding box defaults to Vienna's WGS84 envelope if the caller did
    not provide a ``BOUNDINGBOX_VIENNA`` override.
    """
    bbox = bounding_box or VIENNA_BOUNDING_BOX
    try:
        places = fetch_osm_places()
    except OSMOverpassError as exc:
        # Security (Clear-Text-Logging Drift Round 3): the OSMOverpassError
        # text is itself sanitised at the raise site, but route through
        # ``sanitize_log_arg`` for defence-in-depth — any future shape that
        # surfaces attacker-controlled bytes (e.g. raw protobuf parser
        # excerpts, urllib3 connection-pool stack traces) will be stripped
        # of control characters before reaching the cron runner's stdout.
        logger.error(
            "OSM Overpass enrichment failed: %s",
            sanitize_log_arg(str(exc)),
        )
        return False
    except Exception as exc:  # pragma: no cover - defensive logging path
        # Security (Clear-Text-Logging Drift Round 3): framework catch-all —
        # the exception text is fully attacker-controlled when an upstream
        # library raises a custom subclass with a poisoned ``__str__``.
        logger.error(
            "Unexpected error during OSM Overpass enrichment: %s: %s",
            type(exc).__name__,
            sanitize_log_arg(str(exc)),
        )
        return False

    complete = filter_complete_places(places)
    logger.info(
        "OSM Overpass returned %d candidates (%d after completeness filter)",
        len(places),
        len(complete),
    )
    if not complete:
        return True

    existing_entries = cast(list[StationEntry], [station.as_dict() for station in stations])
    outcome = merge_places(
        existing_entries,
        complete,
        MergeConfig(max_distance_m=merge_distance_m, bounding_box=bbox),
    )

    by_id: dict[str, Mapping[str, object]] = {}
    for entry in outcome.stations:
        bst_id = entry.get("bst_id")
        if isinstance(bst_id, str) and bst_id:
            by_id[str(bst_id)] = entry

    # Tag updated stations with ``source="osm"`` so downstream consumers
    # can tell which directory feed contributed each entry. The merge
    # outcome already records ``source="google_places"`` for stations
    # touched by the Google path; we set ``"osm"`` here without losing
    # any existing source markers.
    updated = 0
    for station in stations:
        merged = by_id.get(station.bst_id)
        if not merged:
            continue
        existing_source = station.extras.get("source")
        sources: set[str] = set()
        if isinstance(existing_source, str):
            sources.update(s.strip() for s in existing_source.split(",") if s.strip())
        sources.add("osm")
        merged_with_source = dict(merged)
        merged_with_source["source"] = ",".join(sorted(sources))
        station.update_from_entry(merged_with_source)
        updated += 1

    logger.info(
        "OSM Overpass enrichment updated %d stations; %d candidates already covered",
        updated,
        len(outcome.skipped_places),
    )
    return True


def _enrich_with_hafas(stations: list[Station]) -> list[Station]:
    """Resolve coordinates for *stations* via the ÖBB HAFAS fallback.

    Iterates over the strict subset of stations still missing
    coordinates after the OSM pass and asks
    :func:`src.places.hafas_client.enrich_station_with_hafas` for each.
    The HAFAS hit is committed straight onto the station's extras —
    matching the field shape produced by the OSM merge — and the
    HAFAS ``extId`` is persisted as the top-level ``hafas_extId`` key
    (parallel to ``vor_id`` / ``_google_place_id``) so downstream
    consumers can resolve a station back to its HAFAS identifier
    without re-querying the upstream.

    Returns the residual list of stations *still* missing coordinates
    after HAFAS ran, so the caller can hand only that subset to the
    Google Places tier — protecting the monthly free-tier quota.
    """
    if not stations:
        return []

    updated = 0
    residual: list[Station] = []
    for station in stations:
        hit = enrich_station_with_hafas(station.name)
        if hit is None:
            residual.append(station)
            continue
        station.extras["latitude"] = hit["lat"]
        station.extras["longitude"] = hit["lon"]
        station.extras["hafas_extId"] = hit["extId"]

        existing_source = station.extras.get("source")
        sources: set[str] = set()
        if isinstance(existing_source, str):
            sources.update(s.strip() for s in existing_source.split(",") if s.strip())
        sources.add("hafas")
        station.extras["source"] = ",".join(sorted(sources))
        updated += 1

    logger.info(
        "HAFAS enrichment resolved coordinates for %d of %d stations; %d still missing",
        updated,
        len(stations),
        len(residual),
    )
    return residual


def _enrich_manual_stations(
    manual_stations: list[dict[str, object]],
    location_index: Mapping[str, LocationInfo],
) -> int:
    """Enrich manual entries (``source=manual``, ``type=manual_*``) with
    coordinates from the local lookup index and the HAFAS LocMatch tier.

    Manual entries — added for the Ostregion Liniennetz stations outside
    the Wien/NÖ pendler scope (PR #1557) — bypass ``_filter_relevant_stations``
    and therefore never enter the ÖBB-side OSM/HAFAS/Google enrichment chain.
    Without this step they stay coordinate-less forever, even though the
    cron pipeline already loads sources (GTFS stops.txt, VOR haltestellen.csv)
    and has an unmetered fallback (HAFAS ÖBB Scotty) that can resolve them.

    Resolution order — same fail-cheap-first rationale as ``_enrich_with_osm``
    / ``_enrich_with_hafas`` / ``_enrich_with_google_places``:

      1. **Local index** (``location_index`` = GTFS + WL + VOR coords,
         free, in-memory). Already built earlier in ``main()`` for the ÖBB
         flag-annotation pass; we just re-use it.
      2. **HAFAS LocMatch** (ÖBB Scotty, free, no monthly quota; the same
         tier that ``_enrich_with_hafas`` taps for ÖBB stations). Handles
         Austria-wide stations plus most CZ/SK/HU border references that
         Scotty knows.

    Skips entries that already carry ``latitude``/``longitude`` so the
    enrichment is idempotent across cron ticks (first run fills 296 entries;
    subsequent runs no-op unless new manual entries appear).

    Returns the number of entries enriched (for log accounting).
    """
    if not manual_stations:
        return 0

    enriched_local = 0
    enriched_hafas = 0
    still_missing = 0
    skipped_existing = 0

    for entry in manual_stations:
        lat = entry.get("latitude")
        lon = entry.get("longitude")
        if isinstance(lat, int | float) and isinstance(lon, int | float):
            skipped_existing += 1
            continue

        name_raw = entry.get("name")
        if not isinstance(name_raw, str) or not name_raw.strip():
            continue
        name = name_raw.strip()

        # 1. Local index. Try the authoritative WL DIVA first (exact key,
        #    immune to the "Wien <name> (WL)" vs "<name>" mismatch), then
        #    fall back to the normalized name keys.
        info: LocationInfo | None = None
        wl_diva = entry.get("wl_diva")
        if isinstance(wl_diva, str) and wl_diva.strip():
            info = location_index.get(_wl_diva_key(wl_diva.strip()))
        if info is None:
            for key in _normalize_location_keys(name):
                info = location_index.get(key)
                if info:
                    break
        if info is not None:
            entry["latitude"] = info.latitude
            entry["longitude"] = info.longitude
            _merge_sources_into_entry(entry, info.sources)
            enriched_local += 1
            continue

        # 2. HAFAS LocMatch fallback (ÖBB Scotty, free)
        try:
            hit = enrich_station_with_hafas(name)
        except Exception:  # nosec B902 - HAFAS must never crash the pipeline
            logger.warning(
                "HAFAS lookup raised for manual entry %s",
                sanitize_log_arg(name),
            )
            still_missing += 1
            continue

        if hit is None:
            still_missing += 1
            continue

        entry["latitude"] = hit["lat"]
        entry["longitude"] = hit["lon"]
        ext_id = hit.get("extId")
        if isinstance(ext_id, str) and ext_id.strip():
            entry["hafas_extId"] = ext_id.strip()
        _merge_sources_into_entry(entry, {"hafas"})
        enriched_hafas += 1

    logger.info(
        "Manual enrichment: %d via local index, %d via HAFAS, %d already had coords, %d still missing",
        enriched_local,
        enriched_hafas,
        skipped_existing,
        still_missing,
    )
    return enriched_local + enriched_hafas


def _merge_sources_into_entry(entry: dict[str, object], add: Iterable[str]) -> None:
    """Add provider tokens to the entry's comma-separated ``source`` field.

    Preserves the existing tokens (typically ``manual``), adds the new ones,
    deduplicates, and re-emits in alphabetical order — matching the
    canonical format pinned in ``src/places/merge.py`` and validated by
    ``_find_naming_issues`` in ``src/utils/stations_validation.py``.
    """
    raw = entry.get("source")
    tokens: set[str] = set()
    if isinstance(raw, str):
        tokens.update(part.strip() for part in raw.split(",") if part.strip())
    elif isinstance(raw, list):
        tokens.update(str(part).strip() for part in raw if str(part).strip())
    for token in add:
        token = str(token).strip()
        if token:
            tokens.add(token)
    entry["source"] = ",".join(sorted(tokens))


def _enrich_with_google_places(
    stations: list[Station],
    *,
    tiles_file: Path | None,
    missing_subset: list[Station] | None = None,
) -> None:
    """Enrich stations via the Google Places fallback.

    *missing_subset* — when provided — is the strict subset of stations
    that the merge step is allowed to touch. The OSM-first / Google-second
    contract requires Google Places to ONLY backfill entries that the
    primary OSM run could not resolve; stations that already carry
    OSM-supplied coordinates must not be re-keyed by the fallback even if
    a Google Place happens to share their name. When *missing_subset* is
    ``None`` the function falls back to the legacy whole-list behaviour
    (used by the no-OSM cron path); callers in the OSM-first flow must
    always pass an explicit subset.
    """
    load_default_env_files()
    env = os.environ

    target_stations = missing_subset if missing_subset is not None else stations
    if not target_stations:
        logger.info("Skipping Google Places enrichment: no stations are missing coordinates")
        return

    try:
        api_key = get_places_api_key()
    except SystemExit as exc:  # pragma: no cover - depends on env configuration
        message = exc.args[0] if exc.args else "Missing GOOGLE_ACCESS_ID"
        logger.warning("Skipping Google Places enrichment: %s", message)
        return

    try:
        tiles = _load_tiles_configuration(tiles_file, env)
    except (OSError, ValueError) as exc:
        logger.error("Cannot load Places tile configuration: %s", exc)
        return

    included_types = _parse_included_types(env.get("PLACES_INCLUDED_TYPES"))
    radius_m = _parse_radius(env.get("PLACES_RADIUS_M"))
    max_result_count = _parse_max_results(env.get("PLACES_MAX_RESULTS"))
    language = env.get("PLACES_LANGUAGE", "de")
    region = env.get("PLACES_REGION", "AT")
    timeout_s = _parse_float(env.get("REQUEST_TIMEOUT_S"), key="REQUEST_TIMEOUT_S", default=25.0)
    max_retries = _parse_int(env.get("REQUEST_MAX_RETRIES"), key="REQUEST_MAX_RETRIES", default=4)
    merge_distance = _parse_float(env.get("MERGE_MAX_DIST_M"), key="MERGE_MAX_DIST_M", default=150.0)

    try:
        bounding_box = _parse_bounding_box(env.get("BOUNDINGBOX_VIENNA"))
    except ValueError as exc:
        logger.error("Invalid BOUNDINGBOX_VIENNA configuration: %s", exc)
        return

    client_config = GooglePlacesConfig(
        api_key=api_key,
        included_types=included_types,
        language=language,
        region=region,
        radius_m=radius_m,
        timeout_s=timeout_s,
        max_retries=max_retries,
        max_result_count=max_result_count,
    )
    client = GooglePlacesClient(client_config)

    try:
        places = _fetch_google_places(client, tiles)
    except GooglePlacesPermissionError as exc:
        hint = permission_hint(str(exc))
        if hint:
            logger.error("Google Places denied access: %s | %s", exc, hint)
        else:
            logger.error("Google Places denied access: %s", exc)
        return
    except GooglePlacesError as exc:
        logger.error("Google Places enrichment failed: %s", exc)
        return

    _merge_google_metadata(
        target_stations,
        places,
        MergeConfig(max_distance_m=merge_distance, bounding_box=bounding_box),
    )


class _NormalizedCSVRow:
    def __init__(self, row: Mapping[str, str | None]):
        self._row = row
        self._map = {_normalize_csv_key(key): key for key in row if key}

    def get(self, *candidates: str) -> str:
        for candidate in candidates:
            key = self._map.get(_normalize_csv_key(candidate))
            if key is None:
                continue
            value = self._row.get(key)
            if value is None:
                continue
            text = str(value).strip()
            if text:
                return text
        return ""


def _iter_vor_rows(path: Path) -> Iterable[_NormalizedCSVRow]:
    # Security: see _load_gtfs_locations / MAX_CSV_LOCATIONS_BYTES for
    # the canonical CSV size-bomb defence shape (read_capped_text +
    # io.StringIO). FileNotFoundError is propagated so the caller's
    # legacy "file not found" log path remains intact.
    if not path.exists():
        raise FileNotFoundError(str(path))
    content = read_capped_text(
        path,
        MAX_CSV_LOCATIONS_BYTES,
        encoding="utf-8-sig",
        label="VOR stops",
        logger=logger,
    )
    if content is None:
        return
    sample = content[:4096]
    delimiter = _detect_csv_delimiter(sample)
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    for row in reader:
        yield _NormalizedCSVRow({key or "": value for key, value in row.items()})


def load_vor_stops(path: Path) -> list[VORStop]:
    try:
        rows = list(_iter_vor_rows(path))
    except FileNotFoundError:
        logger.info(
            "VOR stops file not found: [path-sha256=%s]",
            _path_fingerprint(path),
        )
        return []
    except csv.Error as exc:
        logger.warning(
            "Could not parse VOR stops file [path-sha256=%s]: %s",
            _path_fingerprint(path),
            exc,
        )
        return []

    stops: dict[str, VORStop] = {}
    for row in rows:
        vor_id = row.get(
            "StopPointId",
            "StopID",
            "Stop_Id",
            "StopPoint",
            "ID",
        )
        if not vor_id:
            continue
        name = row.get("StopPointName", "Name", "StopName", "Bezeichnung")
        if not name:
            continue
        municipality = row.get("Municipality", "Gemeinde", "City", "Ort") or None
        short_name = row.get("StopPointShortName", "ShortName", "Kurzname") or None
        stops[vor_id] = VORStop(
            vor_id=vor_id,
            name=name,
            municipality=municipality,
            short_name=short_name,
        )

    if not stops:
        logger.info(
            "No VOR stops extracted from [path-sha256=%s]",
            _path_fingerprint(path),
        )
    else:
        logger.info(
            "Loaded %d VOR stops from [path-sha256=%s]",
            len(stops),
            _path_fingerprint(path),
        )
    return list(stops.values())


def _vor_alias_candidates(stop: VORStop) -> set[str]:
    aliases: set[str] = {stop.name, stop.vor_id}
    if stop.short_name:
        aliases.add(stop.short_name)
    if stop.municipality:
        combined = f"{stop.municipality} {stop.name}".strip()
        aliases.add(combined)
    return {alias for alias in aliases if alias}


def _build_vor_index(stops: Iterable[VORStop]) -> dict[str, list[VORStop]]:
    index: dict[str, list[VORStop]] = {}
    for stop in stops:
        tokens: set[str] = set()
        for alias in _vor_alias_candidates(stop):
            tokens.update(_normalize_location_keys(alias))
        for token in tokens:
            if not token:
                continue
            index.setdefault(token, []).append(stop)
    return index


def _select_vor_stop(station: Station, candidates: list[VORStop]) -> VORStop | None:
    unique: dict[str, VORStop] = {}
    for stop in candidates:
        if stop.vor_id not in unique:
            unique[stop.vor_id] = stop
    stops = list(unique.values())
    if not stops:
        return None
    if len(stops) == 1:
        return stops[0]

    def _is_vienna_stop(stop: VORStop) -> bool:
        return _looks_like_vienna(stop.municipality) or _looks_like_vienna(stop.name)

    if station.in_vienna:
        vienna_stops = [stop for stop in stops if _is_vienna_stop(stop)]
        if len(vienna_stops) == 1:
            return vienna_stops[0]
        if vienna_stops:
            stops = vienna_stops
    else:
        outside = [stop for stop in stops if not _is_vienna_stop(stop)]
        if len(outside) == 1:
            return outside[0]
        if outside:
            stops = outside

    normalized_station = _harmonize_station_name(station.name).casefold()
    name_matches = [stop for stop in stops if _harmonize_station_name(stop.name).casefold() == normalized_station]
    if len(name_matches) == 1:
        return name_matches[0]
    return None


def _load_vor_name_to_id_map(path: Path | None) -> dict[str, str]:
    """Read ``vor-haltestellen.mapping.json`` produced by fetch_vor_haltestellen.

    Returns a dict from the *exact* station name we asked for (i.e. the
    ÖBB Excel-name) to the resolved VOR id. Lets us short-circuit the
    fuzzy `_select_vor_stop` matcher: if the fetcher already determined
    `Hohenau` resolves to `430377800`, we reuse that mapping instead of
    re-deriving it from the candidate list (which can yield ambiguous
    or empty results when the resolved name carries a heavy disambiguation
    suffix like `Hohenau an der March Bahnhof`).
    """
    if path is None or not path.exists():
        return {}
    # Security: ``read_capped_json`` enforces both the depth-bomb catch
    # tuple and the byte-size cap (see MAX_JSON_FILE_BYTES). Same
    # cron-pipeline blast radius as the other loaders in this script.
    payload = read_capped_json(
        path,
        MAX_JSON_FILE_BYTES,
        label="VOR mapping",
    )
    if payload is None:
        logger.warning(
            "Could not read VOR mapping [path-sha256=%s] (missing/invalid/oversized)",
            _path_fingerprint(path),
        )
        return {}
    if not isinstance(payload, list):
        return {}
    mapping: dict[str, str] = {}
    for entry in payload:
        if not isinstance(entry, dict):
            continue
        name = entry.get("station_name")
        vor_id = entry.get("vor_id")
        if isinstance(name, str) and isinstance(vor_id, str):
            text_name = name.strip()
            text_id = vor_id.strip()
            if text_name and text_id:
                mapping[text_name] = text_id
    if mapping:
        logger.info("Loaded %d direct name→vor_id mappings", len(mapping))
    return mapping


def _assign_vor_ids(
    stations: list[Station],
    vor_stops: list[VORStop],
    name_to_vor_id: Mapping[str, str] | None = None,
) -> None:
    name_map = name_to_vor_id or {}
    if not vor_stops and not name_map:
        return
    index = _build_vor_index(vor_stops) if vor_stops else {}
    # Track which vor_ids are already claimed by another station so we
    # never assign the same VOR id twice — that produces the
    # cross_station_id_issues collision the 2026-05 cron exposed
    # (Mistelbach + Mistelbach Stadt both ending up with 430420200).
    # Pre-load existing assignments from `vor_id` already on the
    # stations (bypasses idempotency on re-runs).
    used_vor_ids: set[str] = {station.vor_id for station in stations if station.vor_id}

    def _try_claim(station: Station, vor_id: str) -> bool:
        if vor_id in used_vor_ids:
            logger.warning(
                "Refusing to assign vor_id=%s to %s (bst_id=%s) — already "
                "claimed by another station; the fetcher resolved both "
                "names to the same VOR stop",
                vor_id,
                station.name,
                station.bst_id,
            )
            return False
        station.vor_id = vor_id
        used_vor_ids.add(vor_id)
        return True

    for station in stations:
        if station.vor_id:
            continue
        # Direct lookup via the fetcher-produced mapping.json wins —
        # avoids fuzzy-matching ambiguity when the resolved name differs
        # from the station name (e.g. Hohenau → Hohenau an der March Bahnhof).
        direct = name_map.get(station.name)
        if direct:
            _try_claim(station, direct)
            continue
        if not index:
            continue
        tokens = _normalize_location_keys(station.name)
        if not tokens:
            continue
        candidates: list[VORStop] = []
        for token in tokens:
            candidates.extend(index.get(token, []))
        if not candidates:
            continue
        selected = _select_vor_stop(station, candidates)
        if selected:
            _try_claim(station, selected.vor_id)
        else:
            logger.debug("Ambiguous VOR stop candidates for %s (%s)", station.name, station.bst_id)


def _harmonize_station_names(
    stations: list[Station],
    existing_entries: dict[str, dict[str, object]],
) -> None:
    if not existing_entries:
        for station in stations:
            station.name = _harmonize_station_name(station.name)
        return

    for station in stations:
        try:
            lookup_id = str(int(float(station.bst_id)))
        except (ValueError, TypeError):
            lookup_id = str(station.bst_id)
        existing = existing_entries.get(lookup_id)
        if existing:
            name_raw = existing.get("name")
            if isinstance(name_raw, str) and name_raw.strip():
                canonical = _harmonize_station_name(name_raw)
                if canonical and canonical != station.name:
                    logger.debug(
                        "Using existing name for %s: %s -> %s",
                        station.bst_id,
                        station.name,
                        canonical,
                    )
                station.name = canonical or station.name
                continue
        station.name = _harmonize_station_name(station.name)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download the ÖBB station directory and export a JSON mapping",
    )
    parser.add_argument(
        "--source-url",
        default=DEFAULT_SOURCE_URL,
        help="URL of the Excel workbook to download",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help="Path to the JSON file that will be written",
    )
    parser.add_argument(
        "--pendler",
        type=Path,
        metavar="PATH",
        default=DEFAULT_PENDLER_PATH,
        help="Path to the JSON file containing pendler station IDs",
    )
    parser.add_argument(
        "--pendler-candidates",
        type=Path,
        metavar="PATH",
        default=DEFAULT_PENDLER_CANDIDATES_PATH,
        help=(
            "Path to the name-based pendler whitelist "
            "(default: data/pendler_candidates.json). Stations whose ÖBB name "
            "matches a candidate here are also marked pendler=true."
        ),
    )
    parser.add_argument(
        "--vor-stops",
        type=Path,
        metavar="PATH",
        default=DEFAULT_VOR_STOPS_PATH,
        help="Path to the VOR stop CSV used for VOR_STATION_IDS",
    )
    parser.add_argument(
        "--geonetz-stops",
        type=Path,
        metavar="PATH",
        default=DEFAULT_GEONETZ_STOPS_PATH,
        help=(
            "Path to the compact ÖBB-Infrastruktur GeoNetz stops JSON "
            "(produced by scripts/extract_oebb_geonetz_stops.py from the "
            "upstream GeoNetz_12-2024.zip). Used to attach UIC eva_nr, "
            "IFOPT-ID and postal address to oebb-source stations."
        ),
    )
    parser.add_argument(
        "--gtfs-stops",
        type=Path,
        metavar="PATH",
        default=DEFAULT_GTFS_STOPS_PATH,
        help="Path to a GTFS stops.txt file for coordinate lookup",
    )
    parser.add_argument(
        "--wl-haltepunkte",
        type=Path,
        metavar="PATH",
        default=DEFAULT_WL_HALTEPUNKTE_PATH,
        help="Path to the Wiener Linien haltepunkte CSV for coordinate lookup",
    )
    parser.add_argument(
        "--osm-enrich",
        dest="osm_enrich",
        action=argparse.BooleanOptionalAction,
        default=True,
        help=("Use OpenStreetMap (Overpass API) as the primary station " "directory source (default: enabled)"),
    )
    parser.add_argument(
        "--google-enrich",
        dest="google_enrich",
        action=argparse.BooleanOptionalAction,
        default=True,
        help=("Use Google Places API as a fallback when OSM data is " "missing for a station (default: enabled)"),
    )
    parser.add_argument(
        "--places-tiles-file",
        type=Path,
        metavar="PATH",
        help="Optional JSON file overriding PLACES_TILES for Google enrichment",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="Print progress information during the update run",
    )
    return parser.parse_args()


def configure_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    # Sentinel: route through SafeFormatter so any raw exception text
    # logged via %s in this script is sanitised at the formatter layer.
    # Lazy import to mirror the script's existing fallback shape (the
    # try/except `from src.X import …` block at module top).
    try:
        from src.feed.logging_safe import setup_script_logging
    except ModuleNotFoundError:  # pragma: no cover - fallback for installed-package mode
        from feed.logging_safe import setup_script_logging  # type: ignore[no-redef]
    setup_script_logging(level)
    logging.getLogger("urllib3").setLevel(logging.WARNING)


def download_workbook(
    url: str, cache_path: Path = DEFAULT_CACHED_WORKBOOK_PATH
) -> BytesIO:
    """Download the ÖBB workbook from *url* with a cached-snapshot fallback.

    Mirrors the soft-fail pattern used by the other upstream sources in
    this pipeline:

    * WL OGD (``scripts/update_wl_stations.py:_download_ogd_csv``) —
      atomic-write to the pinned local CSV on success, read the pinned
      CSV on failure.
    * OSM Overpass — circuit-breaker + smoke-check gate, soft-fail
      to Google Places fallback.
    * Google Places — credentials-gated, soft-fail on quota.

    Prior to this commit ÖBB was the only fail-fast source — a
    transient ``data.oebb.at`` outage (CMS migration, CDN issue,
    weekend maintenance) zeroed out the weekly cron tick with no
    recovery path. The new behaviour:

    1. Try the download.
    2. On success, atomic-write the bytes to *cache_path* so the
       next cron tick has a snapshot to fall back to.
    3. On failure, read *cache_path* if it exists (with a warning
       log so the operator sees the upstream went silent).
    4. On failure with no cache, re-raise the original exception —
       there is no station directory to refresh.

    The cache file is committed to the repository by the weekly
    ``update-stations.yml`` auto-commit step (``add_options: "-A"``)
    so the first cron tick after this lands populates the snapshot
    automatically.
    """
    logger.info("Downloading workbook: %s", url)
    try:
        with session_with_retries(USER_AGENT) as session:
            content = fetch_content_safe(session, url, timeout=REQUEST_TIMEOUT)
    except Exception as exc:
        # Security: ``read_capped_bytes`` enforces the canonical
        # TOCTOU-safe size cap (mirrors :func:`read_capped_json` /
        # :func:`read_capped_text` in :mod:`src.utils.files`). The cap
        # is :data:`MAX_CACHED_WORKBOOK_BYTES` — identical to the
        # HTTP-fetch upper bound :data:`src.utils.http.MAX_PAYLOAD_SIZE`,
        # so a cache file larger than what HTTP could have legitimately
        # produced is rejected as tampered. Pre-fix
        # ``cache_path.read_bytes()`` allocated O(file_size) bytes
        # against a planted-huge cache file (compromised CI runner /
        # partial flush + power loss / parallel orchestrator atomic
        # state swap) and raised ``MemoryError`` past the surrounding
        # cron orchestrator. Post-fix the file is treated as missing
        # (``None`` return) and the fall-through error branch re-raises
        # the original upstream ``exc`` — mirroring the pre-fix shape
        # on a missing-cache miss.
        cached_bytes = read_capped_bytes(
            cache_path,
            MAX_CACHED_WORKBOOK_BYTES,
            label="ÖBB workbook cache",
            logger=logger,
        )
        if cached_bytes is not None:
            logger.warning(
                "Failed to download %s (%s); falling back to cached workbook [path-sha256=%s]",
                url,
                exc,
                _path_fingerprint(cache_path),
            )
            return BytesIO(cached_bytes)
        logger.error(
            "Failed to download %s and no cached workbook at %s — the "
            "station directory cannot be refreshed",
            url,
            cache_path,
        )
        raise

    # Atomic-cache the successful download for future fallback. A
    # cache write failure must not break the run that just succeeded;
    # log and continue.
    try:
        _persist_workbook_snapshot(cache_path, content)
    except OSError as exc:  # pragma: no cover - filesystem-dependent
        logger.warning(
            "Could not cache workbook to %s (%s); continuing with the "
            "downloaded bytes",
            cache_path,
            exc,
        )

    return BytesIO(content)


def _persist_workbook_snapshot(cache_path: Path, payload: bytes) -> None:
    """Atomically persist *payload* to *cache_path*.

    Security profile: *payload* is by contract the public ÖBB
    ``Verzeichnis der Verkehrsstationen`` XLSX — free open data, no
    auth, no PII, no secrets. The file is intentionally committed to
    the repository by the weekly ``update-stations.yml`` auto-commit
    step so the next cron tick has a fall-back snapshot if the live
    ``data.oebb.at`` download fails. This mirrors the pinned-CSV
    pattern already established for ``scripts/update_wl_stations.py
    :_download_ogd_csv``, ``data/vor-haltestellen.csv`` and the
    bundled GTFS stops dump. Threading the bytes through a dedicated
    helper (rather than writing the HTTP-response variable directly
    at the call site) keeps the cache writer's intent legible and
    isolates the file-IO from the network-IO data flow.
    """
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    with atomic_write(cache_path, mode="wb", permissions=0o644) as handle:
        handle.write(payload)


def _normalize_header(value: object | None) -> str:
    if value is None:
        return ""
    normalized = str(value).strip().lower()
    for token in (" ", "-", "_"):
        normalized = normalized.replace(token, "")
    return normalized


def _match_required_headers(row: Iterable[object]) -> dict[str, int]:
    normalized = [_normalize_header(cell) for cell in row]
    column_map: dict[str, int] = {}
    for header_field, candidates in HEADER_VARIANTS.items():
        for index, value in enumerate(normalized):
            if any(value == candidate or value.startswith(candidate) for candidate in candidates):
                column_map[header_field] = index
                break
    return column_map


def _find_header_row(rows: Iterable[tuple[object, ...]]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows, start=1):
        if not row:
            continue
        column_map = _match_required_headers(row)
        if len(column_map) == len(HEADER_VARIANTS):
            return index, column_map
    raise ValueError("Could not identify the header row in the workbook")


def _coerce_bst_id(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        digits = value.strip()
        if not digits.isdigit():
            return None
        return digits
    if isinstance(value, int | float):
        return str(int(value))
    return None


def extract_stations(workbook_stream: BytesIO) -> list[Station]:
    # Security: route through the canonical ``validate_zip_archive_safe``
    # helper so the four orthogonal axes (total size, per-entry size,
    # entry count, filename length) are all bounded BEFORE openpyxl
    # parses the workbook. The prior shape ``sum(info.file_size) > 100
    # MiB`` only closed the total-size axis: a malicious ZIP with
    # millions of empty entries (declared ``file_size=0``) trivially
    # passed the total-sum check while inflating ``infolist()`` to
    # millions of ZipInfo objects, OOMing the cron pipeline before any
    # consumer sees a row. The canonical helper closes that gap and
    # pins the auto-discoverable inventory test
    # ``test_no_unbounded_zipfile_zipfile_in_src_or_scripts``.
    try:
        with zipfile.ZipFile(workbook_stream) as archive:
            validate_zip_archive_safe(archive, label="ÖBB workbook")
    except zipfile.BadZipFile as exc:
        raise ValueError("Invalid workbook file") from exc

    workbook_stream.seek(0)
    workbook = openpyxl.load_workbook(workbook_stream, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        header_row_index, column_map = _find_header_row(worksheet.iter_rows(min_row=1, max_row=25, values_only=True))
        logger.debug("Detected header row at index %s", header_row_index)
        stations: list[Station] = []
        seen_ids: set[str] = set()
        for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            if not row:
                continue
            name_cell = row[column_map["name"]] if column_map["name"] < len(row) else None
            code_cell = row[column_map["bst_code"]] if column_map["bst_code"] < len(row) else None
            id_cell = row[column_map["bst_id"]] if column_map["bst_id"] < len(row) else None
            if name_cell is None or code_cell is None:
                continue
            parsed_id = _coerce_bst_id(id_cell)
            if parsed_id is None or parsed_id in seen_ids:
                continue
            station = Station(
                bst_id=parsed_id,
                bst_code=str(code_cell).strip(),
                name=str(name_cell).strip(),
            )
            seen_ids.add(parsed_id)
            stations.append(station)
        stations.sort(key=lambda item: item.bst_id)
        logger.info("Extracted %d stations", len(stations))
        return stations
    finally:
        workbook.close()


def _is_vienna_station(name: str) -> bool:
    text = name.strip()
    if not text.startswith("Wien"):
        return False
    if len(text) == 4:
        return True
    return len(text) > 4 and text[4] in {" ", "-", "/", "("}


# Cache key: ``Station.extras["_in_vienna_basis"]`` — list of three
# elements ``[basis_lat, basis_lon, polygon_result]`` produced by the
# previous ``_resolve_in_vienna_with_cache`` call. Survives across runs
# via ``_restore_existing_metadata`` because it lives in ``extras``
# (not in the base-key skip list). Stored as a JSON-friendly list (not
# a tuple) so JSON load/save round-trips cleanly.
_IN_VIENNA_CACHE_KEY = "_in_vienna_basis"


def _resolve_in_vienna_with_cache(station: Station, info: LocationInfo) -> bool:
    """Compute ``in_vienna`` for ``station``, reusing the cached result
    when the supplied :class:`LocationInfo` coords haven't drifted past
    :data:`STATION_DRIFT_TOLERANCE_METERS` since the last polygon
    evaluation.

    Why: ``_is_point_in_vienna`` runs a ray-casting algorithm against
    the LANDESGRENZEOGD.json multipolygon (~thousands of vertices). At
    several hundred stations per run the aggregate cost is
    non-trivial, and the *result* is stable for any drift below 150 m
    (the city boundary tolerance is much wider than provider drift).
    Reusing the previous run's result when drift is below threshold
    eliminates the recomputation entirely.

    The cache is stored in ``station.extras[_IN_VIENNA_CACHE_KEY]`` as
    ``[basis_lat, basis_lon, polygon_result]`` — JSON-friendly so it
    survives stations.json round-trips. The drift comparison uses the
    canonical ``use_cached_polygon_result`` helper from
    ``src/utils/geo.py``; the basis key is read defensively against
    schema drift (length checks, type checks) and falls back to a
    full polygon evaluation on any malformed cache.
    """
    cache = station.extras.get(_IN_VIENNA_CACHE_KEY)
    cached_lat: float | None = None
    cached_lon: float | None = None
    cached_result: bool | None = None
    if isinstance(cache, list) and len(cache) == 3:
        raw_lat, raw_lon, raw_result = cache
        if isinstance(raw_lat, int | float):
            cached_lat = float(raw_lat)
        if isinstance(raw_lon, int | float):
            cached_lon = float(raw_lon)
        if isinstance(raw_result, bool):
            cached_result = raw_result

    cached = use_cached_polygon_result(
        cached_lat, cached_lon, cached_result, info.latitude, info.longitude
    )
    if cached is not None:
        return cached

    # Cache miss — run the real polygon check and refresh the cache.
    result = bool(_is_point_in_vienna(info.latitude, info.longitude))
    station.extras[_IN_VIENNA_CACHE_KEY] = [
        info.latitude, info.longitude, result
    ]
    return result


def _annotate_station_flags(
    stations: list[Station],
    pendler_ids: set[str],
    locations: Mapping[str, LocationInfo],
    pendler_name_candidates: set[str] | None = None,
) -> None:
    """Set ``in_vienna`` and ``pendler`` mutually exclusively.

    A station is either inside the Vienna city limits *or* a commuter-belt
    station outside; never both. If a Vienna station's bst_id is mistakenly
    listed in ``data/pendler_bst_ids.json``, the ``in_vienna`` flag wins and
    the pendler flag stays ``False`` — a warning is logged so the entry can
    be removed from the whitelist.

    Pendler classification sources (any of them is sufficient):
      1. ``bst_id in pendler_ids`` — legacy bst_id whitelist (`data/pendler_bst_ids.json`).
      2. ``normalized name in pendler_name_candidates`` — name-based wishlist
         (`data/pendler_candidates.json`). Lets the user nominate stations
         without knowing the bst_id; the next ÖBB Excel pull resolves it.
      3. WL-sourced location outside Vienna — auto-promoted (legacy heuristic).
    """
    name_candidates = pendler_name_candidates or set()
    for station in stations:
        info: LocationInfo | None = None
        for key in _normalize_location_keys(station.name):
            info = locations.get(key)
            if info:
                break
        # Resolve in_vienna against the strongest available geo signal.
        # Priority order:
        #   (a) Fresh upstream LocationInfo from this run.
        #   (b) Coords carried forward in ``extras`` by
        #       ``_restore_existing_metadata`` from the prior
        #       stations.json — already validated and inertia-merged.
        #   (c) Name heuristic, last resort.
        # Without (b), a single failed WL/ÖBB/VOR fetch (info=None) for a
        # U-Bahn stop like "Stephansplatz" — whose name does NOT start
        # with "Wien" — would silently flip ``in_vienna`` to False even
        # though the persisted coords clearly fall inside the LANDESGRENZE
        # polygon. The synthetic LocationInfo carries ``sources={"extras"}``
        # so the WL-auto-promote check below remains a no-op for this path.
        polygon_info: LocationInfo | None = info
        if polygon_info is None:
            extras_lat = station.extras.get("latitude")
            extras_lon = station.extras.get("longitude")
            if isinstance(extras_lat, int | float) and isinstance(
                extras_lon, int | float
            ):
                polygon_info = LocationInfo(
                    latitude=float(extras_lat),
                    longitude=float(extras_lon),
                    sources={"extras"},
                )
        if polygon_info is not None:
            # Coordinate Inertia (boundary-check layer): reuse the
            # previous polygon result when info coords haven't drifted
            # past STATION_DRIFT_TOLERANCE_METERS. See
            # ``_resolve_in_vienna_with_cache`` for the full rationale.
            station.in_vienna = _resolve_in_vienna_with_cache(
                station, polygon_info
            )
        else:
            station.in_vienna = _is_vienna_station(station.name)
        pendler_candidate = station.bst_id in pendler_ids
        if not pendler_candidate and name_candidates:
            for key in _normalize_location_keys(station.name):
                if key and key in name_candidates:
                    pendler_candidate = True
                    break
        if info and not station.in_vienna and "wl" in info.sources:
            pendler_candidate = True
        if station.in_vienna and pendler_candidate:
            logger.warning(
                "Station %s (bst_id=%s) is inside Vienna; ignoring pendler "
                "marker — remove the entry from data/pendler_bst_ids.json "
                "or data/pendler_candidates.json",
                station.name,
                station.bst_id,
            )
            pendler_candidate = False
        station.pendler = pendler_candidate


def _filter_relevant_stations(stations: list[Station]) -> list[Station]:
    filtered = [station for station in stations if station.in_vienna or station.pendler]
    removed = len(stations) - len(filtered)
    if removed:
        logger.info("Dropping %d stations outside Vienna and commuter belt", removed)
    return filtered


def _load_geonetz_stops(path: Path) -> dict[str, dict[str, object]]:
    """Read ``data/oebb_geonetz_stops.json`` into a ``bsts_id`` lookup.

    The file is the compact projection produced by
    ``scripts/extract_oebb_geonetz_stops.py`` from the ÖBB-Infrastruktur
    GeoNetz dataset (data.oebb.at GeoNetz_12-2024). Missing file or
    malformed payload degrades to an empty dict with a warning — the
    enrichment is a best-effort metadata-only tier whose absence must
    never crash the cron pipeline.

    Security: ``read_capped_json`` enforces both the depth-bomb catch
    tuple AND the byte-size cap (see :data:`MAX_JSON_FILE_BYTES`),
    plus ``parse_constant`` + ``parse_float`` non-finite literal
    rejection. Pre-fix the loader used ``json.loads(path.read_text(...))``
    which buffers the entire file into memory before parsing — a
    planted huge file at ``data/oebb_geonetz_stops.json`` (compromised
    CI runner / hostile PR / corrupted previous run / partial flush +
    power loss) would propagate ``MemoryError`` (a ``BaseException``
    subclass NOT caught by ``except Exception:``) past the loader and
    crash the weekly station refresh cron pipeline (the orchestrator
    runs every update script via ``subprocess.run(check=True)``).
    Pre-fix also accepted ``NaN`` / ``Infinity`` / scientific-notation
    overflow tokens at the JSON parse boundary — a poisoned coordinate
    field would propagate as ``float('nan')`` / ``float('inf')`` past
    the ``isinstance(value, str)`` guards on ``bsts_id`` into
    downstream enrichment that compares via ``==``/``!=``
    (``nan != nan`` is True — silent dedup invariant breakage).
    Mirrors the canonical loader pattern pinned at
    :func:`_load_existing_station_entries` and
    :func:`load_pendler_station_ids` in this module.
    """
    if not path.exists():
        logger.info("GeoNetz stops file not found: %s — skipping enrichment", path)
        return {}
    # Security: ``read_capped_json`` enforces the depth-bomb catch tuple,
    # the byte-size cap, AND ``parse_constant`` / ``parse_float`` non-
    # finite literal rejection. Returns ``None`` on missing / invalid /
    # oversized files — closing the unbounded ``path.read_text()``
    # MemoryError vector that pre-fix could crash the cron pipeline.
    raw = read_capped_json(
        path,
        MAX_JSON_FILE_BYTES,
        label="GeoNetz stops",
        logger=logger,
    )
    if raw is None:
        logger.warning(
            "Failed to load GeoNetz stops from [path-sha256=%s] "
            "(missing/invalid/oversized)",
            _path_fingerprint(path),
        )
        return {}
    stops = raw.get("stops") if isinstance(raw, Mapping) else None
    if not isinstance(stops, list):
        logger.warning("GeoNetz stops file %s has no 'stops' list", path)
        return {}
    lookup: dict[str, dict[str, object]] = {}
    for entry in stops:
        if not isinstance(entry, Mapping):
            continue
        bsts = entry.get("bsts_id")
        if not isinstance(bsts, str) or not bsts:
            continue
        # Coerce to ``dict[str, object]`` — the loader output uses
        # immutable Mapping interface but downstream want a mutable
        # row they can patch (e.g. for tests). Plain copy is enough.
        lookup[bsts] = dict(entry)
    logger.info("Loaded %d GeoNetz stops from %s", len(lookup), path)
    return lookup


def _enrich_with_geonetz(
    stations: Iterable[Station], geonetz_lookup: Mapping[str, Mapping[str, object]]
) -> None:
    """Attach UIC ``eva_nr``, ``ifopt_id`` and ``address`` from GeoNetz.

    Each station is matched two ways:

    1. **Primary join** by ÖBB ``bst_id`` against the GeoNetz
       ``bsts_id`` column. This covers all 147 oebb-source stations
       cleanly.
    2. **Secondary join** by exact canonical name for entries that
       lack a bst_id match (typically synthetic 900xxx-ids from the
       VOR/WL pathway whose corresponding station nevertheless lives
       in GeoNetz, e.g. Wien Hauptbahnhof, bst_id=900100 in our
       directory vs. BSTS_ID=2393 in GeoNetz). The name match uses
       the unique-name subset of the lookup to avoid false positives
       on the small number of repeated operational names.

    On a hit, the three identifier fields are written to
    ``station.extras`` if not already set. Coordinates are explicitly
    *not* overwritten — they're governed by the OSM/HAFAS/Google
    enrichment tiers and the PR #1601 GeoNetz-reconciliation; this
    loader only adds metadata so the next cron tick doesn't silently
    revert a hand-curated value.

    Idempotent: stations that already carry an ``eva_nr`` from a
    previous run pass through untouched.
    """
    if not geonetz_lookup:
        return
    # Build the secondary by-name index from the same lookup — keep
    # only names that appear exactly once so a duplicate operational
    # name (rare, but happens) can never mis-match.
    name_counts: dict[str, int] = {}
    for row in geonetz_lookup.values():
        nm = row.get("name")
        if isinstance(nm, str) and nm:
            name_counts[nm] = name_counts.get(nm, 0) + 1
    by_name: dict[str, Mapping[str, object]] = {
        str(row["name"]): row
        for row in geonetz_lookup.values()
        if isinstance(row.get("name"), str)
        and name_counts.get(str(row["name"]), 0) == 1
    }
    enriched = 0
    for station in stations:
        geo: Mapping[str, object] | None = None
        bst_id = station.bst_id
        if bst_id:
            try:
                lookup_id = str(int(float(bst_id)))
            except (TypeError, ValueError):
                lookup_id = str(bst_id)
            geo = geonetz_lookup.get(lookup_id)
        if geo is None and station.name:
            geo = by_name.get(station.name)
        if geo is None:
            continue
        changed = False
        for geo_field in ("eva_nr", "ifopt_id", "address"):
            value = geo.get(geo_field)
            if not isinstance(value, str) or not value:
                continue
            if station.extras.get(geo_field):
                continue  # preserve any pre-existing value across cron runs
            station.extras[geo_field] = value
            changed = True
        if changed:
            enriched += 1
            # Append the ``oebb_geonetz`` provenance token to the
            # source list so downstream consumers see the data
            # lineage. Mirrors the convention the PR #1601 coord-
            # reconciliation already set up.
            existing_source = station.extras.get("source")
            tokens: set[str] = set()
            if isinstance(existing_source, str):
                tokens.update(t.strip() for t in existing_source.split(",") if t.strip())
            tokens.add("oebb_geonetz")
            station.extras["source"] = ",".join(sorted(tokens))
    if enriched:
        logger.info("GeoNetz-enriched %d stations with eva_nr/ifopt_id/address", enriched)


def load_pendler_station_ids(path: Path) -> set[str]:
    if not path.exists():
        logger.warning(
            "Pendler station list not found: [path-sha256=%s]",
            _path_fingerprint(path),
        )
        return set()
    # Security: ``read_capped_json`` enforces both the depth-bomb catch
    # tuple and the byte-size cap (see MAX_JSON_FILE_BYTES). When the
    # file exists but is unreadable / invalid / oversized, surface a
    # ``ValueError`` so the canonical exit-1 contract for malformed
    # state is preserved.
    data = read_capped_json(
        path,
        MAX_JSON_FILE_BYTES,
        label="Pendler station list",
    )
    if data is None:
        raise ValueError(f"Invalid JSON in pendler station list: {path}")

    if not isinstance(data, list):
        raise ValueError(f"Pendler station list must be a JSON array: {path}")

    pendler_ids: set[str] = set()
    for entry in data:
        if isinstance(entry, bool):
            raise ValueError(f"Invalid pendler station identifier (boolean) in {path}: {entry!r}")
        if isinstance(entry, int):
            pendler_ids.add(str(entry))
            continue
        if isinstance(entry, str):
            token = entry.strip()
            if token.isdigit():
                pendler_ids.add(token)
                continue
        raise ValueError(f"Invalid pendler station identifier in {path}: {entry!r}")

    logger.info("Loaded %d pendler station IDs", len(pendler_ids))
    return pendler_ids


def load_pendler_name_candidates(path: Path) -> set[str]:
    """Load the name-based pendler whitelist (`data/pendler_candidates.json`).

    Returns a set of normalized name keys (via :func:`_normalize_location_keys`)
    so the caller can do an O(1) ``key in candidates`` check while iterating
    the ÖBB Excel rows. A missing or malformed file degrades to an empty set
    with a warning — the bst_id-based whitelist remains the primary path.

    Both ``name`` and the optional ``alternative_names`` array contribute to
    the key set; this lets the file declare the canonical research-name plus
    the spellings ÖBB itself uses (e.g. "Guntramsdorf Südbahn" canonical with
    "Guntramsdorf" as alternative).
    """
    if not path.exists():
        logger.info(
            "Pendler candidates file not found: [path-sha256=%s] "
            "(using bst_id whitelist only)",
            _path_fingerprint(path),
        )
        return set()
    # Security: ``read_capped_json`` enforces both the depth-bomb catch
    # tuple and the byte-size cap (see MAX_JSON_FILE_BYTES). Same
    # cron-pipeline blast radius as the other loaders in this script.
    data = read_capped_json(
        path,
        MAX_JSON_FILE_BYTES,
        label="Pendler candidates",
    )
    if data is None:
        logger.warning(
            "Invalid JSON in pendler candidates file "
            "[path-sha256=%s] (missing/invalid/oversized)",
            _path_fingerprint(path),
        )
        return set()

    if not isinstance(data, dict):
        logger.warning(
            "Pendler candidates file [path-sha256=%s] must be a JSON object",
            _path_fingerprint(path),
        )
        return set()

    raw = data.get("candidates", [])
    if not isinstance(raw, list):
        logger.warning(
            "Pendler candidates file [path-sha256=%s]: 'candidates' must be a list",
            _path_fingerprint(path),
        )
        return set()

    keys: set[str] = set()
    for entry in raw:
        if not isinstance(entry, dict):
            continue
        name = entry.get("name")
        names: list[str] = []
        if isinstance(name, str) and name.strip():
            names.append(name)
        alternative = entry.get("alternative_names")
        if isinstance(alternative, list):
            for alt in alternative:
                if isinstance(alt, str) and alt.strip():
                    names.append(alt)
        for variant in names:
            for key in _normalize_location_keys(variant):
                if key:
                    keys.add(key)
    logger.info(
        "Loaded %d pendler name-keys from %d candidates",
        len(keys),
        len(raw),
    )
    return keys


def write_json(stations_list: list[dict[str, object]], output_path: Path) -> None:
    # Security (Trojan-Source / BiDi-Mark Drift Round 14, ingestion-boundary
    # defence): strip the canonical CVE-2021-42574 attack-byte union from
    # the incoming stations BEFORE ``json.dump``. The OEBB
    # ``Verzeichnis der Verkehrsstationen`` Excel response is the
    # primary delivery vector — a hijacked OGD portal (DNS rebind or
    # cache poisoning) could plant U+202E in a station ``name`` /
    # ``aliases[]`` / ``alternative_names`` field. The weekly
    # ``update-stations.yml`` cron commits ``data/stations.json`` to
    # ``main``. ``ensure_ascii=False`` preserves compact German diffs.
    # Mirrors ``src/places/merge.py:write_stations`` (Round 13).
    #
    # Security (Coordinate finite/range drift, companion-writer
    # defence-in-depth): ``allow_nan=False`` mirrors the canonical
    # writer-side pin established in Round 1485 at
    # ``src/places/merge.py:write_stations``. The local
    # ``_coerce_float_value`` parser (line 455) accepts
    # ``float('nan')`` / ``float('inf')`` from a compromised OGD
    # Wien / GTFS / VOR upstream because the
    # ``isinstance(value, int | float)`` shape check on JSON-decoded
    # numeric values accepts the non-finite literals that
    # ``json.loads`` parses by default. Without this writer-side
    # floor a poisoned upstream silently lands non-standard ``NaN``
    # / ``Infinity`` literals (invalid per RFC 8259) in the public
    # ``data/stations.json`` artefact.
    scrubbed = scrub_trojan_source_primitives(stations_list)
    serialisable = scrubbed if isinstance(scrubbed, list) else stations_list
    payload = {"stations": serialisable}
    # Use atomic_write to prevent partial writes and reduce race conditions.
    with atomic_write(output_path, mode="w", encoding="utf-8", permissions=0o644) as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2, allow_nan=False)
        handle.write("\n")
    logger.info("Wrote [path-sha256=%s]", _path_fingerprint(output_path))


def main() -> None:
    args = parse_args()
    configure_logging(args.verbose)
    _refresh_provider_caches()
    existing_entries, manual_stations = _load_existing_station_entries(args.output)
    workbook_stream = download_workbook(args.source_url)
    stations = extract_stations(workbook_stream)
    _harmonize_station_names(stations, existing_entries)
    _restore_existing_metadata(stations, existing_entries)
    pendler_ids = load_pendler_station_ids(path=args.pendler)
    pendler_name_candidates = load_pendler_name_candidates(path=args.pendler_candidates)
    location_index = _build_location_index(
        args.gtfs_stops,
        args.wl_haltepunkte,
        vor_path=args.vor_stops,
    )
    if not location_index:
        logger.warning("No coordinate data available; falling back to name heuristic")
    _annotate_station_flags(
        stations,
        pendler_ids,
        location_index,
        pendler_name_candidates=pendler_name_candidates,
    )
    vor_stops = load_vor_stops(args.vor_stops) if args.vor_stops else []
    vor_name_map: dict[str, str] = {}
    if args.vor_stops:
        # vor-haltestellen.mapping.json sits next to the .csv (same stem,
        # ``.mapping.json`` suffix) and is produced by fetch_vor_haltestellen.
        mapping_candidate = args.vor_stops.with_suffix(".mapping.json")
        vor_name_map = _load_vor_name_to_id_map(mapping_candidate)
    if vor_stops or vor_name_map:
        _assign_vor_ids(stations, vor_stops, name_to_vor_id=vor_name_map)
    stations = _filter_relevant_stations(stations)

    # GeoNetz metadata enrichment (PR β). Runs after the filter so it
    # only touches the pendler/in_vienna subset, and before OSM/HAFAS/
    # Google so downstream enrichment tiers see the eva_nr that PR γ
    # (planned: HAFAS-drift-detection) needs as the join key. Idempotent
    # — re-runs leave previously-set values alone.
    geonetz_lookup = _load_geonetz_stops(args.geonetz_stops)
    if geonetz_lookup:
        _enrich_with_geonetz(stations, geonetz_lookup)

    # OSM is now the primary directory enrichment source. Google Places
    # only runs as a *fallback* when at least one station is still
    # missing coordinates after OSM completed (or when OSM itself
    # failed). This keeps the Google Places monthly free-tier quota
    # untouched when the public Overpass API already covers the entire
    # directory.
    #
    # ``WIEN_OEPNV_OSM_ENRICH=0`` env-disables OSM without taking the
    # CLI flag — used by the wrapper-orchestrator regression test
    # (``tests/test_update_all_stations_wrapper.py:test_wrapper_atomic_on_success``)
    # to keep its 60-second pytest-timeout budget free of real Overpass
    # round-trips. Production cron runs leave the env unset, so OSM
    # remains the primary source by default.
    load_default_env_files()
    env = os.environ
    cli_enabled = bool(getattr(args, "osm_enrich", False))
    env_enabled = get_bool_env("WIEN_OEPNV_OSM_ENRICH", True)
    osm_succeeded = False
    if cli_enabled and env_enabled:
        try:
            bounding_box = _parse_bounding_box(env.get("BOUNDINGBOX_VIENNA"))
        except ValueError as exc:
            logger.error("Invalid BOUNDINGBOX_VIENNA configuration: %s", exc)
            bounding_box = None
        merge_distance = _parse_float(env.get("MERGE_MAX_DIST_M"), key="MERGE_MAX_DIST_M", default=150.0)
        osm_succeeded = _enrich_with_osm(
            stations,
            bounding_box=bounding_box,
            merge_distance_m=merge_distance,
        )
    elif not cli_enabled:
        logger.info("Skipping OSM Overpass enrichment (--no-osm-enrich)")
    else:
        logger.info("Skipping OSM Overpass enrichment (WIEN_OEPNV_OSM_ENRICH=0)")

    if getattr(args, "google_enrich", False):
        missing = _stations_missing_coordinates(stations)
        if not missing:
            logger.info(
                "Skipping Google Places enrichment: %s already covered all " "%d stations with coordinates",
                "OSM" if osm_succeeded else "the existing directory",
                len(stations),
            )
        else:
            # HAFAS (ÖBB Scotty) is the second-tier fallback: it sits
            # between the OSM Overpass primary and the Google Places
            # last-resort. The HAFAS Mgate API has no per-month quota,
            # so resolving coordinates here directly reduces the load
            # on the Google free-tier budget tracked in
            # data/places_quota.json.
            missing = _enrich_with_hafas(missing)
            if not missing:
                logger.info(
                    "HAFAS resolved every remaining station; skipping Google Places enrichment"
                )
            else:
                if osm_succeeded:
                    logger.info(
                        "Falling back to Google Places for %d stations still " "missing coordinates after OSM + HAFAS enrichment",
                        len(missing),
                    )
                else:
                    logger.info(
                        "Falling back to Google Places for %d stations missing " "coordinates (OSM Overpass was unavailable)",
                        len(missing),
                    )
                # Pass the strict subset so Google never re-keys stations that
                # OSM / HAFAS (or any earlier source) already resolved. The merge
                # logic in src.places.merge would otherwise greedily match Google
                # Places by name even when the existing entry is complete, giving
                # the fallback authority it shouldn't have.
                _enrich_with_google_places(
                    stations,
                    tiles_file=args.places_tiles_file,
                    missing_subset=missing,
                )
    else:
        logger.info("Skipping Google Places enrichment (--no-google-enrich)")

    # Enrich the manual block (manual_distant_at / manual_foreign_city —
    # the Ostregion Liniennetz stations from PR #1557) that bypassed the
    # ÖBB filter and therefore the ÖBB-side enrichment chain. Re-uses
    # the location_index built earlier (free GTFS/WL/VOR lookup) and falls
    # back to the unmetered HAFAS LocMatch tier — exactly the same
    # cheap-first strategy the ÖBB pipeline already runs.
    #
    # Env-disabled via ``WIEN_OEPNV_MANUAL_ENRICH=0`` — mirrors the
    # ``WIEN_OEPNV_OSM_ENRICH`` gate used by
    # ``test_wrapper_atomic_on_success`` to keep the orchestrator
    # subprocess test under its 180-second pytest timeout: the test
    # carries 296 manual entries without coordinates, and 296 real
    # HAFAS LocMatch round-trips from a GitHub-hosted runner regularly
    # burn 3-5 minutes (well over the budget). Production cron runs
    # leave the env unset so the enrichment remains active.
    if get_bool_env("WIEN_OEPNV_MANUAL_ENRICH", True):
        _enrich_manual_stations(manual_stations, location_index)
    else:
        logger.info(
            "Skipping manual entry enrichment (WIEN_OEPNV_MANUAL_ENRICH=0)"
        )

    final_stations = [station.as_dict() for station in stations]
    final_stations.extend(manual_stations)
    write_json(final_stations, args.output)


if __name__ == "__main__":
    main()
